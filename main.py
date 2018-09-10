
import tweepy
from textblob import TextBlob
import csv
from array import *
import json
import re
import nltk
from nltk import pos_tag, word_tokenize
import openpyxl
from openpyxl import Workbook

consumer_key = 'Your Key here'
consumer_secret = 'Your Key here'

access_token = 'Your key here'
access_token_secret ='Your Key here'

auth = tweepy.OAuthHandler(consumer_key,consumer_secret)
auth.set_access_token(access_token,access_token_secret)

api = tweepy.API(auth)
keys = ['tweets','polarity']
tweetings=[]
sentiments=[]
csvdata = []
mentions = []
mentionindex = []
alltweet = []
# alltweet = api.user_timeline(screen_name='anikethgireesh',count=10)
book = openpyxl.load_workbook('F:\games\Machine Learning\datasettwi.xlsx')

sheet = book.active

book2 = Workbook()

sheet2 = book2.active
readat=[]

temp=[]
temp2=[]
coun1=0
coun2=0
coun3=0
b=[]
c=[]
d=[]
for row in sheet.iter_rows(min_row=6, min_col=1, max_row=300, max_col=2):
	oe=1
	for cell in row:
		if oe == 1:
			oe=2
			try:
				tweet = api.get_status(cell.value)
				alltweet.append(tweet.text)
				a = []
				try:
					temp=(tweet.text).split(" ")
					
					for wor in temp:
						print(wor)
						if('@' in wor):
							coun1+=1
						if('#' in wor):
							coun2+=1
						if('http' in wor):
							coun3+=1
					a.append(coun1)
					b.append(coun2)
					c.append(coun3)
					d.append(coun1)
					d.append(coun2)
					d.append(coun3)
					d.append()	
					print(coun1, coun2, coun3)
					coun1=0
					coun2=0
					coun3=0
				except:
					a.append("0")
					b.append("0")
					c.append("0")
					coun1=0
					coun2=0
					coun3=0
					pass
				a[:] = []
				b[:] = []
				c[:] = []
				print(tweet.text, end=" ")
				
			except:
				pass
		else:
			oe=1
			d.append(cell.value)
			sheet2.append(d)
			book2.save('database2re.xlsx')
			d[:]=[]

	print()
